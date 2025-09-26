from openpyxl import Workbook

# Crear un nuevo workbook
wb = Workbook()
ws = wb.active
ws.title = "Contactos"

# Agregar encabezados
ws['A1'] = 'nombre'
ws['B1'] = 'apellido'
ws['C1'] = 'numero'
ws['D1'] = 'carrera'
ws['E1'] = 'grupo'

# Datos de ejemplo
datos = [
    ['Juan', 'Pérez', '51987654321', 'Ingeniería de Sistemas', 'Ingeniería'],
    ['María', 'González', '51912345678', 'Medicina', 'Medicina'],
    ['Carlos', 'López', '51911223344', 'Derecho', 'Derecho'],
    ['Ana', 'Martínez', '51999887766', 'Psicología', 'Psicología'],
    ['Luis', 'Rodríguez', '51955443322', 'Administración', 'Administración'],
    ['Sofia', 'Fernández', '51977665544', 'Arquitectura', 'Ingeniería'],
    ['Diego', 'García', '51933445566', 'Contabilidad', 'Administración'],
    ['Valentina', 'Hernández', '51988990011', 'Enfermería', 'Medicina'],
    ['Miguel', 'Jiménez', '51922334455', 'Marketing', 'Administración'],
    ['Camila', 'Morales', '51966778899', 'Educación', 'Psicología']
]

# Agregar datos
for i, fila in enumerate(datos, start=2):
    ws[f'A{i}'] = fila[0]
    ws[f'B{i}'] = fila[1]
    ws[f'C{i}'] = fila[2]
    ws[f'D{i}'] = fila[3]
    ws[f'E{i}'] = fila[4]

# Guardar archivo
wb.save('ejemplo_contactos.xlsx')

print("✅ Archivo 'ejemplo_contactos.xlsx' creado exitosamente")
print("📋 Contiene 10 contactos de ejemplo con los campos requeridos")
print("🚀 Puedes usar este archivo para probar la importación")
