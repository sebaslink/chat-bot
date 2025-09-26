#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Crear archivo Excel de ejemplo completo con todos los grupos implementados
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def crear_excel_completo():
    """Crear archivo Excel completo con grupos"""
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="25D366", end_color="25D366", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Agregar encabezados con estilo
    encabezados = ['nombre', 'apellido', 'numero', 'carrera', 'grupo']
    for i, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=i, value=encabezado)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos de ejemplo más completos
    datos = [
        # Ingeniería
        ['Juan', 'Pérez', '51987654321', 'Ingeniería de Sistemas', 'Ingeniería'],
        ['Sofia', 'Fernández', '51977665544', 'Arquitectura', 'Ingeniería'],
        ['Diego', 'García', '51933445566', 'Ingeniería Civil', 'Ingeniería'],
        ['Carlos', 'Mendoza', '51944556677', 'Ingeniería Industrial', 'Ingeniería'],
        
        # Medicina
        ['María', 'González', '51912345678', 'Medicina', 'Medicina'],
        ['Valentina', 'Hernández', '51988990011', 'Enfermería', 'Medicina'],
        ['Ana', 'Rodríguez', '51911223344', 'Odontología', 'Medicina'],
        ['Luis', 'Vargas', '51955667788', 'Farmacia', 'Medicina'],
        
        # Derecho
        ['Carlos', 'López', '51911223344', 'Derecho', 'Derecho'],
        ['Patricia', 'Silva', '51999887766', 'Derecho Penal', 'Derecho'],
        ['Roberto', 'Castro', '51933445566', 'Derecho Civil', 'Derecho'],
        
        # Administración
        ['Luis', 'Rodríguez', '51955443322', 'Administración', 'Administración'],
        ['Miguel', 'Jiménez', '51922334455', 'Marketing', 'Administración'],
        ['Carmen', 'Torres', '51966778899', 'Contabilidad', 'Administración'],
        ['Fernando', 'Ramos', '51988990011', 'Recursos Humanos', 'Administración'],
        
        # Psicología
        ['Ana', 'Martínez', '51999887766', 'Psicología', 'Psicología'],
        ['Camila', 'Morales', '51966778899', 'Educación', 'Psicología'],
        ['Isabel', 'Gutiérrez', '51944556677', 'Psicología Clínica', 'Psicología'],
        ['Pedro', 'Díaz', '51911223344', 'Psicología Organizacional', 'Psicología'],
        
        # General (sin grupo específico)
        ['Alejandro', 'Vega', '51933445566', 'Comunicación Social', 'General'],
        ['Laura', 'Cruz', '51955667788', 'Arte', 'General']
    ]
    
    # Agregar datos
    for i, fila in enumerate(datos, start=2):
        for j, valor in enumerate(fila, 1):
            ws.cell(row=i, column=j, value=valor)
    
    # Ajustar ancho de columnas
    column_widths = [15, 15, 15, 25, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Agregar información adicional en una hoja separada
    ws_info = wb.create_sheet("Instrucciones")
    ws_info['A1'] = "INSTRUCCIONES PARA IMPORTAR CONTACTOS"
    ws_info['A1'].font = Font(bold=True, size=14)
    
    instrucciones = [
        "",
        "1. Este archivo contiene contactos de ejemplo con grupos predefinidos.",
        "",
        "2. Grupos disponibles:",
        "   - Ingeniería: Estudiantes y profesionales de ingeniería",
        "   - Medicina: Estudiantes y profesionales de medicina",
        "   - Derecho: Estudiantes y profesionales de derecho", 
        "   - Administración: Estudiantes y profesionales de administración",
        "   - Psicología: Estudiantes y profesionales de psicología",
        "   - General: Grupo por defecto",
        "",
        "3. Columnas requeridas:",
        "   - nombre: Nombre del contacto (obligatorio)",
        "   - apellido: Apellido del contacto (opcional)",
        "   - numero: Número de teléfono (obligatorio, solo números)",
        "   - carrera: Carrera o profesión (opcional)",
        "   - grupo: Grupo al que pertenece (opcional, se mapea automáticamente)",
        "",
        "4. Para importar:",
        "   - Ve a la pestaña 'Importar' en la aplicación",
        "   - Selecciona este archivo Excel",
        "   - Los grupos se asignarán automáticamente",
        "",
        "5. Notas:",
        "   - Los números deben incluir código de país (ej: 51987654321)",
        "   - Los grupos deben coincidir exactamente con los nombres predefinidos",
        "   - Se pueden agregar más contactos siguiendo el mismo formato"
    ]
    
    for i, instruccion in enumerate(instrucciones, 2):
        ws_info[f'A{i}'] = instruccion
    
    # Ajustar ancho de columna en hoja de instrucciones
    ws_info.column_dimensions['A'].width = 60
    
    # Guardar archivo
    wb.save('ejemplo_contactos_completo.xlsx')
    
    print("✅ Archivo 'ejemplo_contactos_completo.xlsx' creado exitosamente")
    print("📋 Contiene 20 contactos de ejemplo con todos los grupos")
    print("📖 Incluye hoja de instrucciones detalladas")
    print("🚀 Listo para importar en el sistema")

if __name__ == '__main__':
    crear_excel_completo()
