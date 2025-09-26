#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Crear plantilla Excel mejorada con columna de grupos bien destacada
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crear_plantilla_mejorada():
    """Crear plantilla Excel mejorada con columna de grupos"""
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="25D366", end_color="25D366", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Estilo para la columna de grupos (destacada)
    grupo_font = Font(bold=True, color="FFFFFF", size=12)
    grupo_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    
    # Borde
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    encabezados = [
        ('A1', 'nombre', 'Nombre del contacto'),
        ('B1', 'apellido', 'Apellido del contacto'),
        ('C1', 'numero', 'Número de teléfono'),
        ('D1', 'carrera', 'Carrera o profesión'),
        ('E1', 'grupo', 'Grupo (IMPORTANTE)')
    ]
    
    # Aplicar estilos a encabezados
    for cell_ref, encabezado, descripcion in encabezados:
        cell = ws[cell_ref]
        cell.value = encabezado
        
        if encabezado == 'grupo':
            # Destacar la columna de grupos
            cell.font = grupo_font
            cell.fill = grupo_fill
        else:
            cell.font = header_font
            cell.fill = header_fill
            
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos de ejemplo con grupos
    datos = [
        ['Juan', 'Pérez', '51987654321', 'Ingeniería de Sistemas', 'Ingeniería'],
        ['María', 'González', '51912345678', 'Medicina', 'Medicina'],
        ['Carlos', 'López', '51911223344', 'Derecho', 'Derecho'],
        ['Ana', 'Martínez', '51999887766', 'Psicología', 'Psicología'],
        ['Luis', 'Rodríguez', '51955443322', 'Administración', 'Administración'],
        ['Sofia', 'Fernández', '51977665544', 'Arquitectura', 'Ingeniería'],
        ['Diego', 'García', '51933445566', 'Contabilidad', 'Administración'],
        ['Valentina', 'Hernández', '51988990011', 'Enfermería', 'Medicina'],
        ['Miguel', 'Jiménez', '51922334455', 'Marketing', 'Administración'],
        ['Camila', 'Morales', '51966778899', 'Educación', 'Psicología']
    ]
    
    # Agregar datos con estilos
    for i, fila in enumerate(datos, start=2):
        for j, valor in enumerate(fila, 1):
            cell = ws.cell(row=i, column=j, value=valor)
            cell.border = thin_border
            
            # Destacar la columna de grupos
            if j == 5:  # Columna E (grupo)
                cell.font = Font(bold=True, color="FF6B35")
    
    # Ajustar ancho de columnas
    column_widths = [15, 15, 15, 25, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Agregar información sobre grupos en la parte inferior
    ws['A12'] = "GRUPOS DISPONIBLES:"
    ws['A12'].font = Font(bold=True, size=12, color="25D366")
    
    grupos_info = [
        ("A13", "Ingeniería", "Estudiantes y profesionales de ingeniería"),
        ("A14", "Medicina", "Estudiantes y profesionales de medicina"),
        ("A15", "Derecho", "Estudiantes y profesionales de derecho"),
        ("A16", "Administración", "Estudiantes y profesionales de administración"),
        ("A17", "Psicología", "Estudiantes y profesionales de psicología"),
        ("A18", "General", "Grupo por defecto (sin especificar)")
    ]
    
    for cell_ref, grupo, descripcion in grupos_info:
        ws[cell_ref] = f"{grupo}: {descripcion}"
        ws[cell_ref].font = Font(size=10)
    
    # Agregar instrucciones
    ws['A20'] = "INSTRUCCIONES:"
    ws['A20'].font = Font(bold=True, size=12, color="25D366")
    
    instrucciones = [
        "1. Completa todas las filas con los datos de tus contactos",
        "2. En la columna 'grupo', usa EXACTAMENTE uno de los grupos disponibles arriba",
        "3. Los números deben incluir código de país (ej: 51987654321)",
        "4. Guarda el archivo y súbelo en la aplicación",
        "5. Los grupos se asignarán automáticamente a tus contactos"
    ]
    
    for i, instruccion in enumerate(instrucciones, 21):
        ws[f'A{i}'] = instruccion
        ws[f'A{i}'].font = Font(size=10)
    
    # Crear hoja de instrucciones detalladas
    ws_inst = wb.create_sheet("Instrucciones Detalladas")
    
    # Título
    ws_inst['A1'] = "GUÍA COMPLETA PARA IMPORTAR CONTACTOS CON GRUPOS"
    ws_inst['A1'].font = Font(bold=True, size=16, color="25D366")
    
    # Contenido de instrucciones
    contenido = [
        "",
        "PASO 1: PREPARAR TUS DATOS",
        "• Asegúrate de tener la siguiente información para cada contacto:",
        "  - Nombre (obligatorio)",
        "  - Apellido (opcional)",
        "  - Número de teléfono con código de país (obligatorio)",
        "  - Carrera o profesión (opcional)",
        "  - Grupo al que pertenece (opcional pero recomendado)",
        "",
        "PASO 2: USAR LA COLUMNA DE GRUPOS",
        "• La columna 'grupo' es muy importante para organizar tus contactos",
        "• Usa EXACTAMENTE uno de estos nombres de grupo:",
        "  ✓ Ingeniería",
        "  ✓ Medicina", 
        "  ✓ Derecho",
        "  ✓ Administración",
        "  ✓ Psicología",
        "  ✓ General",
        "",
        "PASO 3: COMPLETAR EL ARCHIVO",
        "• Copia la estructura de las filas de ejemplo",
        "• Reemplaza los datos de ejemplo con tus contactos reales",
        "• Mantén el mismo formato de columnas",
        "• No cambies los nombres de las columnas",
        "",
        "PASO 4: IMPORTAR EN LA APLICACIÓN",
        "• Ve a la pestaña 'Importar' en la aplicación",
        "• Selecciona este archivo Excel",
        "• Los grupos se asignarán automáticamente",
        "• Podrás enviar mensajes por grupo específico",
        "",
        "BENEFICIOS DE USAR GRUPOS:",
        "• Organización mejorada de contactos",
        "• Envío de mensajes dirigidos por área profesional",
        "• Filtrado eficiente de contactos",
        "• Gestión centralizada de diferentes audiencias",
        "",
        "NOTAS IMPORTANTES:",
        "• Los números deben tener formato internacional (ej: 51987654321)",
        "• Los nombres de grupos deben coincidir exactamente",
        "• Puedes dejar la columna grupo vacía (se asignará 'General')",
        "• El sistema validará los datos antes de importar"
    ]
    
    for i, texto in enumerate(contenido, 2):
        cell = ws_inst[f'A{i}']
        cell.value = texto
        if texto.startswith(('PASO', 'BENEFICIOS', 'NOTAS')):
            cell.font = Font(bold=True, size=12, color="25D366")
        elif texto.startswith('✓'):
            cell.font = Font(bold=True, color="25D366")
        else:
            cell.font = Font(size=10)
    
    # Ajustar ancho de columna en hoja de instrucciones
    ws_inst.column_dimensions['A'].width = 80
    
    # Guardar archivo
    wb.save('plantilla_contactos_con_grupos_mejorada.xlsx')
    
    print("✅ Plantilla Excel mejorada creada exitosamente")
    print("📋 Archivo: plantilla_contactos_con_grupos_mejorada.xlsx")
    print("🎯 Características:")
    print("   • Columna 'grupo' destacada en color naranja")
    print("   • Lista de grupos disponibles en la hoja")
    print("   • Instrucciones detalladas incluidas")
    print("   • Hoja separada con guía completa")
    print("   • Ejemplos claros de uso")

if __name__ == '__main__':
    crear_plantilla_mejorada()
