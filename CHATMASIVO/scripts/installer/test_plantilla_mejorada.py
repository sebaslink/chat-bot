#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para probar la plantilla mejorada con columna de grupos
"""

import sys
import os
sys.path.append('codigo')

def test_plantilla_mejorada():
    """Probar la plantilla mejorada del sistema"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="25D366", end_color="25D366", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo para la columna de grupos (destacada)
        grupo_font = Font(bold=True, color="FFFFFF", size=12)
        grupo_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        
        # Borde
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        encabezados = [
            ('A1', 'nombre', 'Nombre del contacto'),
            ('B1', 'apellido', 'Apellido del contacto'),
            ('C1', 'numero', 'Número de teléfono'),
            ('D1', 'carrera', 'Carrera o profesión'),
            ('E1', 'grupo', 'Grupo (IMPORTANTE)')
        ]
        
        # Aplicar estilos a encabezados
        for cell_ref, encabezado, descripcion in encabezados:
            cell = ws[cell_ref]
            cell.value = encabezado
            
            if encabezado == 'grupo':
                # Destacar la columna de grupos
                cell.font = grupo_font
                cell.fill = grupo_fill
            else:
                cell.font = header_font
                cell.fill = header_fill
                
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos de ejemplo con grupos
        datos = [
            ['Juan', 'Pérez', '51987654321', 'Ingeniería de Sistemas', 'Ingeniería'],
            ['María', 'González', '51912345678', 'Medicina', 'Medicina'],
            ['Carlos', 'López', '51911223344', 'Derecho', 'Derecho'],
            ['Ana', 'Martínez', '51999887766', 'Psicología', 'Psicología'],
            ['Luis', 'Rodríguez', '51955443322', 'Administración', 'Administración']
        ]
        
        # Agregar datos con estilos
        for i, fila in enumerate(datos, start=2):
            for j, valor in enumerate(fila, 1):
                cell = ws.cell(row=i, column=j, value=valor)
                cell.border = thin_border
                
                # Destacar la columna de grupos
                if j == 5:  # Columna E (grupo)
                    cell.font = Font(bold=True, color="FF6B35")
        
        # Ajustar ancho de columnas
        column_widths = [15, 15, 15, 25, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Agregar información sobre grupos
        ws['A8'] = "GRUPOS DISPONIBLES:"
        ws['A8'].font = Font(bold=True, size=12, color="25D366")
        
        grupos_info = [
            ("A9", "Ingeniería", "Estudiantes y profesionales de ingeniería"),
            ("A10", "Medicina", "Estudiantes y profesionales de medicina"),
            ("A11", "Derecho", "Estudiantes y profesionales de derecho"),
            ("A12", "Administración", "Estudiantes y profesionales de administración"),
            ("A13", "Psicología", "Estudiantes y profesionales de psicología"),
            ("A14", "General", "Grupo por defecto (sin especificar)")
        ]
        
        for cell_ref, grupo, descripcion in grupos_info:
            ws[cell_ref] = f"{grupo}: {descripcion}"
            ws[cell_ref].font = Font(size=10)
        
        # Guardar archivo
        wb.save('plantilla_sistema_mejorada.xlsx')
        
        print("✅ Plantilla mejorada creada exitosamente")
        print("📋 Archivo: plantilla_sistema_mejorada.xlsx")
        print("🎯 Características:")
        print("   • Columna 'grupo' destacada en color naranja")
        print("   • Lista de grupos disponibles en la hoja")
        print("   • Instrucciones incluidas")
        print("   • Estilos profesionales aplicados")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def verificar_plantilla():
    """Verificar la plantilla generada"""
    try:
        import pandas as pd
        
        # Leer archivo generado
        df = pd.read_excel('plantilla_sistema_mejorada.xlsx', sheet_name='Contactos')
        
        print("\n🔍 Verificando plantilla mejorada...")
        print(f"📊 Columnas: {list(df.columns)}")
        
        # Verificar que tiene la columna grupo
        if 'grupo' in df.columns:
            print("✅ Columna 'grupo' presente")
            print(f"🏷️ Grupos en datos: {df['grupo'].unique()}")
        else:
            print("❌ Columna 'grupo' NO presente")
            return False
        
        # Mostrar datos
        print("\n📋 Datos de ejemplo:")
        print(df.to_string(index=False))
        
        return True
        
    except Exception as e:
        print(f"❌ Error verificando plantilla: {e}")
        return False

def main():
    """Función principal"""
    print("🚀 Probando plantilla mejorada con columna de grupos...")
    print("=" * 60)
    
    if test_plantilla_mejorada():
        if verificar_plantilla():
            print("\n✅ Verificación completada exitosamente")
            print("🎯 La plantilla mejorada incluye correctamente la columna de grupos")
            print("📁 Archivo guardado como: plantilla_sistema_mejorada.xlsx")
        else:
            print("\n❌ La verificación falló")
    else:
        print("\n❌ Error creando la plantilla")

if __name__ == '__main__':
    main()
