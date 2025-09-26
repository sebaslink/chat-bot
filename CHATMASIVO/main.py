from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
import sqlite3
import os
import logging
from twilio.rest import Client
from dotenv import load_dotenv
import random
from datetime import datetime, timedelta
import csv
import io
from werkzeug.utils import secure_filename
import base64

# Integración del chatbot deshabilitada
CHATBOT_INTEGRATION_AVAILABLE = False

load_dotenv('config/twilio/Twilio.env')

app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'tu_clave_secreta_aqui')

# Configuración para subir archivos
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Crear directorio de uploads si no existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Verificar si el archivo tiene una extensión permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_extension(filename):
    """Obtener la extensión del archivo"""
    return filename.rsplit('.', 1)[1].lower()

def crear_plantilla_csv():
    """Crear plantilla CSV como alternativa a Excel"""
    try:
        import csv
        import io
        
        # Crear contenido CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Encabezados
        writer.writerow(['nombre', 'apellido', 'numero', 'carrera', 'grupo'])
        
        # Datos de ejemplo con grupos H.S y facultades
        datos = [
            ['Juan', 'Pérez', '51987654321', 'Ingeniería de Sistemas', 'Ingenierías'],
            ['María', 'González', '51912345678', 'Medicina', 'Medicina'],
            ['Carlos', 'López', '51911223344', 'Derecho', 'Derecho'],
            ['Ana', 'Martínez', '51999887766', 'Psicología', 'Psicología'],
            ['Luis', 'Rodríguez', '51955443322', 'Administración', 'Administración'],
            ['Sofia', 'Fernández', '51977665544', 'Arquitectura', 'Ingenierías'],
            ['Diego', 'García', '51933445566', 'Contabilidad', 'Administración'],
            ['Valentina', 'Hernández', '51988990011', 'Enfermería', 'Medicina'],
            ['Miguel', 'Jiménez', '51922334455', 'Marketing', 'Administración'],
            ['Camila', 'Morales', '51966778899', 'Educación', 'Psicología'],
            ['Pedro', 'Sánchez', '51944556677', 'Estudiante de Secundaria', 'H.S'],
            ['Laura', 'Vega', '51955667788', 'Estudiante de Secundaria', 'H.S']
        ]
        
        # Agregar datos
        for fila in datos:
            writer.writerow(fila)
        
        # Agregar información sobre grupos
        writer.writerow([])  # Línea vacía
        writer.writerow(['GRUPOS DISPONIBLES:'])
        writer.writerow(['H.S', 'Estudiantes de High School (Secundaria)'])
        writer.writerow(['Ingenierías', 'Estudiantes y profesionales de ingeniería'])
        writer.writerow(['Medicina', 'Estudiantes y profesionales de medicina'])
        writer.writerow(['Derecho', 'Estudiantes y profesionales de derecho'])
        writer.writerow(['Administración', 'Estudiantes y profesionales de administración'])
        writer.writerow(['Psicología', 'Estudiantes y profesionales de psicología'])
        writer.writerow(['General', 'Grupo por defecto (sin especificar)'])
        
        # Agregar instrucciones
        writer.writerow([])  # Línea vacía
        writer.writerow(['INSTRUCCIONES:'])
        writer.writerow(['1. Completa todas las filas con los datos de tus contactos'])
        writer.writerow(['2. En la columna "grupo", usa EXACTAMENTE uno de los grupos disponibles arriba'])
        writer.writerow(['3. Los números deben incluir código de país (ej: 51987654321)'])
        writer.writerow(['4. Guarda el archivo y súbelo en la aplicación'])
        writer.writerow(['5. Los grupos se asignarán automáticamente a tus contactos'])
        
        # Convertir a bytes
        csv_content = output.getvalue()
        output.close()
        
        return app.response_class(
            csv_content,
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=plantilla_contactos_con_grupos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'}
        )
        
    except Exception as e:
        logger.error(f"Error creando plantilla CSV: {e}")
        return jsonify({'success': False, 'message': f'Error creando plantilla: {str(e)}'})

def get_mime_type(extension):
    """Obtener el tipo MIME basado en la extensión"""
    mime_types = {
        'png': 'image/png',
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg'
    }
    return mime_types.get(extension, 'image/jpeg')

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('chatmasivo.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuración Twilio
TWILIO_ACCOUNT_SID = os.getenv('TWILIO_ACCOUNT_SID')
TWILIO_AUTH_TOKEN = os.getenv('TWILIO_AUTH_TOKEN')
TWILIO_WHATSAPP_FROM = os.getenv('TWILIO_WHATSAPP_FROM', '+13073182623')

# Número de prueba
NUMERO_PRUEBA = "+15005550009"

# Inicializar cliente Twilio
try:
    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
except:
    client = None

# Integración del chatbot deshabilitada
chatbot_integration = None

def get_db_connection():
    """Obtener conexión a la base de datos creando el directorio si es necesario"""
    os.makedirs('data/database', exist_ok=True)
    return sqlite3.connect('data/database/numeros_whatsapp.db')

def init_db():
    """Crear base de datos y tablas si no existen"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Tabla de grupos
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS grupos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE,
            descripcion TEXT,
            activo BOOLEAN DEFAULT 1,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Tabla de números (actualizada)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS numeros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            apellido TEXT,
            telefono TEXT NOT NULL UNIQUE,
            carrera TEXT,
            grupo_id INTEGER DEFAULT NULL,
            activo BOOLEAN DEFAULT 1,
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (grupo_id) REFERENCES grupos (id)
        )
    ''')
    
    # Tabla de logs de mensajes
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mensajes_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_id INTEGER,
            mensaje TEXT,
            status TEXT,
            twilio_sid TEXT,
            error_message TEXT,
            fecha_envio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (numero_id) REFERENCES numeros (id)
        )
    ''')
    
    # Tabla de plantillas de mensajes
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS plantillas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            intro TEXT,
            texto_random TEXT,
            texto_fijo TEXT,
            activa BOOLEAN DEFAULT 1,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Insertar grupos por defecto
    grupos_default = [
        ('H.S', 'Estudiantes de High School (Secundaria)'),
        ('General', 'Grupo por defecto para todos los contactos'),
        ('Ingenierías', 'Estudiantes y profesionales de ingeniería'),
        ('Medicina', 'Estudiantes y profesionales de medicina'),
        ('Derecho', 'Estudiantes y profesionales de derecho'),
        ('Administración', 'Estudiantes y profesionales de administración'),
        ('Psicología', 'Estudiantes y profesionales de psicología')
    ]
    
    for nombre, descripcion in grupos_default:
        cursor.execute('SELECT COUNT(*) FROM grupos WHERE nombre = ?', (nombre,))
        if cursor.fetchone()[0] == 0:
            cursor.execute('INSERT INTO grupos (nombre, descripcion) VALUES (?, ?)', 
                          (nombre, descripcion))
    
    conn.commit()
    conn.close()
    logger.info("Base de datos inicializada correctamente")

def get_numeros_activos(grupo_id=None):
    """Obtener todos los números activos, opcionalmente filtrados por grupo"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if grupo_id:
        cursor.execute('''
            SELECT n.id, n.nombre, n.apellido, n.telefono, n.carrera, g.nombre as grupo_nombre 
            FROM numeros n 
            LEFT JOIN grupos g ON n.grupo_id = g.id 
            WHERE n.activo = 1 AND n.grupo_id = ?
        ''', (grupo_id,))
    else:
        cursor.execute('''
            SELECT n.id, n.nombre, n.apellido, n.telefono, n.carrera, g.nombre as grupo_nombre 
            FROM numeros n 
            LEFT JOIN grupos g ON n.grupo_id = g.id 
            WHERE n.activo = 1
        ''')
    
    numeros = cursor.fetchall()
    conn.close()
    return numeros

def get_grupos():
    """Obtener todos los grupos activos"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, nombre, descripcion FROM grupos')
    grupos = cursor.fetchall()
    conn.close()
    return grupos

def crear_grupo(nombre, descripcion=""):
    """Crear nuevo grupo"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('INSERT INTO grupos (nombre, descripcion) VALUES (?, ?)', (nombre, descripcion))
        conn.commit()
        conn.close()
        logger.info(f"Grupo '{nombre}' creado exitosamente")
        return True
    except sqlite3.IntegrityError:
        logger.warning(f"Grupo '{nombre}' ya existe")
        return False

def guardar_plantilla(nombre, intro, texto_random, texto_fijo):
    """Guardar plantilla de mensaje"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO plantillas (nombre, intro, texto_random, texto_fijo) 
            VALUES (?, ?, ?, ?)
        ''', (nombre, intro, texto_random, texto_fijo))
        conn.commit()
        conn.close()
        logger.info(f"Plantilla '{nombre}' guardada exitosamente")
        return True
    except Exception as e:
        logger.error(f"Error al guardar plantilla: {e}")
        return False

def get_plantillas():
    """Obtener todas las plantillas activas"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, nombre, intro, texto_random, texto_fijo FROM plantillas WHERE activa = 1')
    plantillas = cursor.fetchall()
    conn.close()
    return plantillas

def log_mensaje(numero_id, mensaje, status, twilio_sid=None, error_message=None):
    """Registrar envío de mensaje en el log"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener el número de teléfono si tenemos numero_id
        numero_telefono = None
        if numero_id:
            cursor.execute('SELECT telefono FROM numeros WHERE id = ?', (numero_id,))
            result = cursor.fetchone()
            if result:
                numero_telefono = result[0]
        
        cursor.execute('''
            INSERT INTO mensajes_log (numero_id, mensaje, status, twilio_sid, error_message) 
            VALUES (?, ?, ?, ?, ?)
        ''', (numero_id, mensaje, status, twilio_sid, error_message))
        conn.commit()
        conn.close()
        logger.info(f"Mensaje loggeado: {status} para número ID {numero_id}")
    except Exception as e:
        logger.error(f"Error loggeando mensaje: {e}")

def get_estadisticas():
    """Obtener estadísticas de envíos"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Total de contactos
    cursor.execute('SELECT COUNT(*) FROM numeros WHERE activo = 1')
    total_contactos = cursor.fetchone()[0]
    
    # Mensajes enviados hoy
    cursor.execute('''
        SELECT COUNT(*) FROM mensajes_log 
        WHERE status = 'enviado' AND DATE(fecha_envio) = DATE('now')
    ''')
    mensajes_hoy = cursor.fetchone()[0]
    
    # Mensajes enviados esta semana
    cursor.execute('''
        SELECT COUNT(*) FROM mensajes_log 
        WHERE status = 'enviado' AND fecha_envio >= datetime('now', '-7 days')
    ''')
    mensajes_semana = cursor.fetchone()[0]
    
    # Errores recientes
    cursor.execute('''
        SELECT COUNT(*) FROM mensajes_log 
        WHERE status = 'error' AND fecha_envio >= datetime('now', '-24 hours')
    ''')
    errores_recientes = cursor.fetchone()[0]
    
    conn.close()
    
    return {
        'total_contactos': total_contactos,
        'mensajes_hoy': mensajes_hoy,
        'mensajes_semana': mensajes_semana,
        'errores_recientes': errores_recientes
    }

def agregar_numero(nombre, apellido, telefono, carrera, grupo_id=None):
    """Agregar nuevo número a la base de datos"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('INSERT INTO numeros (nombre, apellido, telefono, carrera, grupo_id) VALUES (?, ?, ?, ?, ?)', 
                      (nombre, apellido, telefono, carrera, grupo_id))
        conn.commit()
        conn.close()
        logger.info(f"Número {telefono} agregado exitosamente")
        return True
    except sqlite3.IntegrityError:
        logger.warning(f"El número {telefono} ya existe")
        return False  # Número ya existe

def desactivar_numero(numero_id):
    """Desactivar número (no eliminar)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('UPDATE numeros SET activo = 0 WHERE id = ?', (numero_id,))
    conn.commit()
    conn.close()

def eliminar_numero(numero_id):
    """Eliminar número completamente de la base de datos"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener información del número antes de eliminar
        cursor.execute('SELECT nombre, apellido, telefono FROM numeros WHERE id = ?', (numero_id,))
        numero_info = cursor.fetchone()
        
        if not numero_info:
            conn.close()
            return False, "Número no encontrado"
        
        # Eliminar el número
        cursor.execute('DELETE FROM numeros WHERE id = ?', (numero_id,))
        
        # Eliminar logs de mensajes relacionados
        cursor.execute('DELETE FROM mensajes_log WHERE numero_id = ?', (numero_id,))
        
        conn.commit()
        conn.close()
        
        logger.info(f"Número eliminado: {numero_info[0]} {numero_info[1]} ({numero_info[2]})")
        return True, f"Contacto {numero_info[0]} {numero_info[1]} eliminado exitosamente"
        
    except Exception as e:
        logger.error(f"Error eliminando número {numero_id}: {e}")
        return False, f"Error eliminando contacto: {str(e)}"

def eliminar_todos_contactos():
    """Eliminar todos los contactos de la base de datos"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Contar contactos antes de eliminar
        cursor.execute('SELECT COUNT(*) FROM numeros')
        total_contactos = cursor.fetchone()[0]
        
        # Eliminar todos los contactos
        cursor.execute('DELETE FROM numeros')
        
        # Limpiar logs de mensajes también
        cursor.execute('DELETE FROM mensajes_log')
        
        conn.commit()
        conn.close()
        
        logger.info(f"Todos los contactos eliminados: {total_contactos} contactos")
        return True, f"Se eliminaron {total_contactos} contactos exitosamente"
        
    except Exception as e:
        logger.error(f"Error eliminando contactos: {e}")
        return False, f"Error eliminando contactos: {str(e)}"

def generar_mensaje(nombre, intro="mane", texto_random=None, texto_fijo="Este es el mensaje predeterminado."):
    """Generar mensaje personalizado con integración del chatbot"""
    opciones_random = [
        "Tenemos una novedad que podría interesarte.",
        "Te comparto una actualización rápida.",
        "Solo paso a dejarte este aviso importante.",
        "Gracias por tu atención, aquí va la info."
    ]
    
    if not texto_random:
        texto_random = random.choice(opciones_random)
    
    # Integración del chatbot deshabilitada
    
    # Mensaje tradicional si no hay chatbot o falló
    partes = [
        intro.strip(),
        f"Hola {nombre},".strip(),
        texto_random.strip(),
        texto_fijo.strip()
    ]
    return "\n\n".join(p for p in partes if p)

def enviar_whatsapp_masivo(intro="mane", texto_fijo="Este es el mensaje predeterminado.", 
                          texto_random=None, grupo_id=None, imagen_path=None):
    """Enviar mensajes masivos por WhatsApp con soporte para imágenes"""
    if not client:
        logger.error("Cliente Twilio no configurado")
        return {"success": False, "message": "Cliente Twilio no configurado"}
    
    numeros = get_numeros_activos(grupo_id)
    resultados = []
    
    logger.info(f"Iniciando envío masivo a {len(numeros)} contactos")
    
    for numero_id, nombre, apellido, telefono, carrera, grupo_nombre in numeros:
        try:
            mensaje = generar_mensaje(nombre, intro, texto_random, texto_fijo)
            
            # Formatear número para WhatsApp (E.164)
            if not telefono.startswith('+'):
                telefono = '+' + telefono
            if not telefono.startswith('whatsapp:'):
                telefono = 'whatsapp:' + telefono
            
            # Preparar parámetros del mensaje
            message_params = {
                'from_': TWILIO_WHATSAPP_FROM,
                'to': f'whatsapp:{telefono}',
                'body': mensaje
            }
            
            # Agregar imagen si existe
            if imagen_path and os.path.exists(imagen_path):
                # Leer imagen y convertir a base64
                with open(imagen_path, 'rb') as img_file:
                    img_data = img_file.read()
                    img_base64 = base64.b64encode(img_data).decode('utf-8')
                    
                    # Obtener tipo MIME
                    extension = get_file_extension(imagen_path)
                    mime_type = get_mime_type(extension)
                    
                    # Agregar media al mensaje
                    message_params['media_url'] = [f"data:{mime_type};base64,{img_base64}"]
                    logger.info(f"Imagen agregada al mensaje para {nombre}")
            
            # Enviar mensaje
            message = client.messages.create(**message_params)
            
            # Log del mensaje exitoso
            log_mensaje(numero_id, mensaje, "enviado", message.sid)
            
            resultados.append({
                "nombre": f"{nombre} {apellido or ''}".strip(),
                "telefono": telefono,
                "carrera": carrera,
                "grupo": grupo_nombre,
                "status": "enviado",
                "sid": message.sid,
                "con_imagen": bool(imagen_path and os.path.exists(imagen_path))
            })
            
            logger.info(f"Mensaje enviado exitosamente a {nombre} ({telefono})")
            
        except Exception as e:
            # Log del error
            log_mensaje(numero_id, mensaje if 'mensaje' in locals() else "", "error", None, str(e))
            
            resultados.append({
                "nombre": f"{nombre} {apellido or ''}".strip(),
                "telefono": telefono,
                "carrera": carrera,
                "grupo": grupo_nombre,
                "status": "error",
                "error": str(e)
            })
            
            logger.error(f"Error enviando mensaje a {nombre} ({telefono}): {e}")
    
    logger.info(f"Envío masivo completado. Exitosos: {len([r for r in resultados if r['status'] == 'enviado'])}, Errores: {len([r for r in resultados if r['status'] == 'error'])}")
    return {"success": True, "resultados": resultados}

@app.route('/api/user-info')
def get_user_info():
    """Obtener información del usuario desde el sistema principal"""
    try:
        import requests
        # Hacer una petición al sistema principal para obtener la información del usuario
        response = requests.get('http://localhost:5000/api/auth/check', 
                              cookies=request.cookies, 
                              timeout=5)
        if response.status_code == 200:
            data = response.json()
            # Asegurar que la respuesta tenga el formato correcto
            if 'authenticated' in data and data['authenticated']:
                return data
            else:
                return {"authenticated": False, "message": "No autenticado"}
        else:
            return {"authenticated": False, "message": "No autenticado"}
    except requests.exceptions.ConnectionError:
        return {"authenticated": False, "message": "Sistema principal no disponible"}
    except requests.exceptions.Timeout:
        return {"authenticated": False, "message": "Tiempo de espera agotado"}
    except Exception as e:
        logger.error(f"Error obteniendo información del usuario: {e}")
        return {"authenticated": False, "message": f"Error: {str(e)}"}

@app.route('/')
def index():
    """Página principal"""
    numeros = get_numeros_activos()
    grupos = get_grupos()
    plantillas = get_plantillas()
    estadisticas = get_estadisticas()
    return render_template('chatmasivo_original.html', 
                         numeros=numeros, 
                         grupos=grupos,
                         plantillas=plantillas,
                         estadisticas=estadisticas,
                         numero_prueba=NUMERO_PRUEBA)

@app.route('/agregar', methods=['POST'])
def agregar():
    """Agregar nuevo número"""
    nombre = request.form.get('nombre')
    apellido = request.form.get('apellido', '')
    telefono = request.form.get('telefono')
    carrera = request.form.get('carrera', '')
    grupo_id = request.form.get('grupo_id')
    
    if not nombre or not telefono:
        flash('Nombre y teléfono son requeridos', 'error')
        return redirect(url_for('index'))
    
    # Limpiar número (quitar espacios, guiones, etc.)
    telefono = ''.join(filter(str.isdigit, telefono))
    
    # Convertir grupo_id a int si existe
    grupo_id = int(grupo_id) if grupo_id and grupo_id != '' else None
    
    if agregar_numero(nombre, apellido, telefono, carrera, grupo_id):
        flash(f'Número {telefono} agregado exitosamente', 'success')
    else:
        flash(f'El número {telefono} ya existe', 'error')
    
    return redirect(url_for('index'))

@app.route('/desactivar/<int:numero_id>')
def desactivar(numero_id):
    """Desactivar número"""
    desactivar_numero(numero_id)
    flash('Número desactivado', 'info')
    return redirect(url_for('index'))

@app.route('/eliminar/<int:numero_id>')
def eliminar(numero_id):
    """Eliminar número completamente"""
    try:
        success, message = eliminar_numero(numero_id)
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/eliminar_todos_contactos', methods=['POST'])
def eliminar_todos_contactos_route():
    """Eliminar todos los contactos"""
    try:
        success, message = eliminar_todos_contactos()
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/enviar_masivo', methods=['POST'])
def enviar_masivo():
    """Enviar mensajes masivos con soporte para imágenes"""
    intro = request.form.get('intro', 'mane')
    texto_fijo = request.form.get('texto_fijo', 'Este es el mensaje predeterminado.')
    texto_random = request.form.get('texto_random', '')
    grupo_id = request.form.get('grupo_id', '')
    
    if not texto_random:
        texto_random = None
    
    # Convertir grupo_id a int si existe
    grupo_id = int(grupo_id) if grupo_id and grupo_id != '' else None
    
    # Manejar imagen si se subió
    imagen_path = None
    if 'imagen' in request.files:
        archivo_imagen = request.files['imagen']
        if archivo_imagen and archivo_imagen.filename != '':
            if allowed_file(archivo_imagen.filename):
                # Generar nombre único para el archivo
                filename = secure_filename(archivo_imagen.filename)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                name, ext = os.path.splitext(filename)
                filename = f"{name}_{timestamp}{ext}"
                
                # Guardar archivo
                imagen_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                archivo_imagen.save(imagen_path)
                logger.info(f"Imagen guardada: {imagen_path}")
            else:
                flash('Formato de imagen no válido. Use PNG, JPG o JPEG', 'error')
                return redirect(url_for('index'))
    
    resultado = enviar_whatsapp_masivo(intro, texto_fijo, texto_random, grupo_id, imagen_path)
    
    # Limpiar imagen temporal después del envío
    if imagen_path and os.path.exists(imagen_path):
        try:
            os.remove(imagen_path)
            logger.info(f"Imagen temporal eliminada: {imagen_path}")
        except Exception as e:
            logger.warning(f"No se pudo eliminar imagen temporal: {e}")
    
    if resultado['success']:
        enviados = len([r for r in resultado['resultados'] if r['status'] == 'enviado'])
        errores = len([r for r in resultado['resultados'] if r['status'] == 'error'])
        con_imagen = len([r for r in resultado['resultados'] if r.get('con_imagen', False)])
        
        mensaje = f'Mensajes enviados: {enviados}, Errores: {errores}'
        if con_imagen > 0:
            mensaje += f', Con imagen: {con_imagen}'
        flash(mensaje, 'success')
    else:
        flash(f'Error: {resultado["message"]}', 'error')
    
    return redirect(url_for('index'))

@app.route('/api/enviar_prueba', methods=['POST'])
def enviar_prueba():
    """API para enviar mensaje de prueba"""
    if not client:
        return jsonify({"success": False, "message": "Cliente Twilio no configurado"})
    
    try:
        mensaje = generar_mensaje("Usuario de Prueba", "mane", None, "Mensaje de prueba desde el sistema.")
        
        message = client.messages.create(
            from_=TWILIO_WHATSAPP_FROM,
            to=f'whatsapp:{NUMERO_PRUEBA}',
            body=mensaje
        )
        
        return jsonify({
            "success": True, 
            "message": "Mensaje de prueba enviado",
            "sid": message.sid
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/crear_grupo', methods=['POST'])
def crear_grupo_route():
    """Crear nuevo grupo"""
    nombre = request.form.get('nombre')
    descripcion = request.form.get('descripcion', '')
    
    if not nombre:
        flash('Nombre del grupo es requerido', 'error')
        return redirect(url_for('index'))
    
    if crear_grupo(nombre, descripcion):
        flash(f'Grupo "{nombre}" creado exitosamente', 'success')
    else:
        flash(f'El grupo "{nombre}" ya existe', 'error')
    
    return redirect(url_for('index'))

@app.route('/guardar_plantilla', methods=['POST'])
def guardar_plantilla_route():
    """Guardar plantilla de mensaje"""
    nombre = request.form.get('nombre')
    intro = request.form.get('intro', 'mane')
    texto_random = request.form.get('texto_random', '')
    texto_fijo = request.form.get('texto_fijo', '')
    
    if not nombre or not texto_fijo:
        flash('Nombre y texto fijo son requeridos', 'error')
        return redirect(url_for('index'))
    
    if guardar_plantilla(nombre, intro, texto_random, texto_fijo):
        flash(f'Plantilla "{nombre}" guardada exitosamente', 'success')
    else:
        flash(f'Error al guardar plantilla', 'error')
    
    return redirect(url_for('index'))

@app.route('/exportar_contactos')
def exportar_contactos():
    """Exportar contactos a CSV"""
    numeros = get_numeros_activos()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Nombre', 'Apellido', 'Teléfono', 'Carrera', 'Grupo'])
    
    for numero in numeros:
        writer.writerow([numero[0], numero[1], numero[2] or '', numero[3], numero[4] or '', numero[5] or 'Sin grupo'])
    
    output.seek(0)
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'contactos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )

@app.route('/importar_excel', methods=['POST'])
def importar_excel():
    """Importar contactos desde Excel"""
    try:
        if 'archivo_excel' not in request.files:
            flash('No se seleccionó archivo', 'error')
            return redirect(url_for('index'))
        
        archivo = request.files['archivo_excel']
        grupo_id = request.form.get('grupo_importar', '')
        
        if archivo.filename == '':
            flash('No se seleccionó archivo', 'error')
            return redirect(url_for('index'))
        
        if archivo and archivo.filename.endswith(('.xlsx', '.xls')):
            # Convertir grupo_id a int si existe
            grupo_id = int(grupo_id) if grupo_id and grupo_id != '' else None
            
            # Procesar archivo
            resultado = importar_desde_excel(archivo, grupo_id)
            
            if resultado['success']:
                res = resultado['resultados']
                mensaje = f'Importación completada: {res["exitosos"]} exitosos, {res["errores"]} errores, {res["duplicados"]} duplicados'
                if res.get('limite_alcanzado'):
                    mensaje += ' (Límite de 5000 contactos alcanzado)'
                flash(mensaje, 'success')
                
                # Mostrar detalles en logs
                for detalle in res['detalles'][:10]:  # Mostrar solo los primeros 10
                    logger.info(detalle)
            else:
                flash(f'Error: {resultado["message"]}', 'error')
        else:
            flash('Formato de archivo no válido. Use .xlsx o .xls', 'error')
            
    except Exception as e:
        logger.error(f"Error en importación: {e}")
        flash(f'Error procesando archivo: {str(e)}', 'error')
    
    return redirect(url_for('index'))

def importar_desde_excel(archivo_excel, grupo_id=None):
    """Importar contactos desde archivo Excel con límite de 5000 y manejo de bloqueos"""
    try:
        import pandas as pd
        import time
        
        # Leer el archivo Excel
        df = pd.read_excel(archivo_excel)
        
        # Verificar límite antes de procesar
        contactos_actuales = len(get_numeros_activos())
        if contactos_actuales >= 5000:
            return {
                'success': False,
                'message': f'Límite de 5000 contactos alcanzado. Contactos actuales: {contactos_actuales}'
            }
        
        # Verificar que tenga las columnas necesarias
        columnas_requeridas = ['nombre', 'apellido', 'numero', 'carrera', 'grupo']
        columnas_disponibles = [col.lower().strip() for col in df.columns]
        
        # Mapear columnas (case insensitive)
        mapeo_columnas = {}
        for col_req in columnas_requeridas:
            for col_disp in columnas_disponibles:
                if col_req in col_disp or col_disp in col_req:
                    mapeo_columnas[col_req] = df.columns[columnas_disponibles.index(col_disp)]
                    break
        
        # Verificar que se encontraron todas las columnas
        if len(mapeo_columnas) < 5:
            return {
                'success': False, 
                'message': f'Faltan columnas requeridas. Se encontraron: {list(mapeo_columnas.keys())}'
            }
        
        # Procesar cada fila
        resultados = {
            'exitosos': 0,
            'errores': 0,
            'duplicados': 0,
            'limite_alcanzado': False,
            'detalles': []
        }
        
        for index, row in df.iterrows():
            # Verificar límite en cada iteración
            if len(get_numeros_activos()) >= 5000:
                resultados['limite_alcanzado'] = True
                resultados['detalles'].append(f"Límite de 5000 contactos alcanzado en fila {index + 2}")
                break
                
            try:
                nombre = str(row[mapeo_columnas['nombre']]).strip()
                apellido = str(row[mapeo_columnas['apellido']]).strip()
                numero = str(row[mapeo_columnas['numero']]).strip()
                carrera = str(row[mapeo_columnas['carrera']]).strip()
                grupo_nombre = str(row[mapeo_columnas['grupo']]).strip()
                
                # Limpiar número (solo dígitos)
                numero_limpio = ''.join(filter(str.isdigit, numero))
                
                if not nombre or not numero_limpio:
                    resultados['errores'] += 1
                    resultados['detalles'].append(f"Fila {index + 2}: Faltan datos obligatorios")
                    continue
                
                # Obtener ID del grupo por nombre con reintentos
                grupo_id_final = grupo_id  # Usar el grupo_id pasado como parámetro si existe
                if not grupo_id_final and grupo_nombre:
                    grupo_id_final = buscar_grupo_por_nombre_con_reintentos(grupo_nombre)
                
                # Intentar agregar el contacto con reintentos
                success = agregar_numero_con_reintentos(nombre, apellido, numero_limpio, carrera, grupo_id_final)
                if success:
                    resultados['exitosos'] += 1
                    resultados['detalles'].append(f"Fila {index + 2}: {nombre} {apellido} - Agregado exitosamente")
                else:
                    resultados['duplicados'] += 1
                    resultados['detalles'].append(f"Fila {index + 2}: {nombre} {apellido} - Número duplicado")
                    
            except Exception as e:
                resultados['errores'] += 1
                resultados['detalles'].append(f"Fila {index + 2}: Error - {str(e)}")
        
        logger.info(f"Importación desde Excel completada: {resultados['exitosos']} exitosos, {resultados['errores']} errores")
        return {'success': True, 'resultados': resultados}
        
    except Exception as e:
        logger.error(f"Error en importación Excel: {e}")
        return {'success': False, 'message': str(e)}

def buscar_grupo_por_nombre_con_reintentos(nombre_grupo):
    """Buscar grupo por nombre con manejo de bloqueos"""
    max_intentos = 3
    for intento in range(max_intentos):
        try:
            conn = sqlite3.connect('data/database/numeros_whatsapp.db', timeout=10)
            cursor = conn.cursor()
            cursor.execute('SELECT id FROM grupos WHERE nombre = ?', (nombre_grupo,))
            grupo_result = cursor.fetchone()
            conn.close()
            return grupo_result[0] if grupo_result else None
        except sqlite3.OperationalError as e:
            if "database is locked" in str(e) and intento < max_intentos - 1:
                time.sleep(0.5)  # Esperar 500ms antes de reintentar
                continue
            else:
                logger.error(f"Error buscando grupo: {e}")
                return None
        except Exception as e:
            logger.error(f"Error inesperado buscando grupo: {e}")
            return None
    return None

def agregar_numero_con_reintentos(nombre, apellido, telefono, carrera, grupo_id):
    """Agregar número con manejo de bloqueos"""
    max_intentos = 3
    for intento in range(max_intentos):
        try:
            conn = sqlite3.connect('data/database/numeros_whatsapp.db', timeout=10)
            cursor = conn.cursor()
            
            # Verificar si ya existe
            cursor.execute('SELECT id FROM numeros WHERE telefono = ?', (telefono,))
            if cursor.fetchone():
                conn.close()
                return False  # Ya existe
            
            # Insertar nuevo contacto
            cursor.execute('INSERT INTO numeros (nombre, apellido, telefono, carrera, grupo_id) VALUES (?, ?, ?, ?, ?)',
                          (nombre, apellido, telefono, carrera, grupo_id))
            conn.commit()
            conn.close()
            
            logger.info(f"Número {telefono} agregado exitosamente")
            return True
            
        except sqlite3.OperationalError as e:
            if "database is locked" in str(e) and intento < max_intentos - 1:
                time.sleep(0.5)
                continue
            else:
                logger.error(f"Error agregando número: {e}")
                return False
        except Exception as e:
            logger.error(f"Error inesperado agregando número: {e}")
            return False
    
    return False

@app.route('/descargar_plantilla_excel')
def descargar_plantilla_excel():
    """Descargar plantilla de Excel para importar contactos con grupos"""
    try:
        # Intentar usar openpyxl si está disponible
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            openpyxl_available = True
        except ImportError:
            openpyxl_available = False
            print("Advertencia: openpyxl no está disponible, creando plantilla CSV")
        
        if openpyxl_available:
            # Crear workbook Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Contactos"
        else:
            # Crear plantilla CSV como alternativa
            return crear_plantilla_csv()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="25D366", end_color="25D366", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo para la columna de grupos (destacada)
        grupo_font = Font(bold=True, color="FFFFFF", size=12)
        grupo_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        
        # Borde
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        encabezados = [
            ('A1', 'nombre', 'Nombre del contacto'),
            ('B1', 'apellido', 'Apellido del contacto'),
            ('C1', 'numero', 'Número de teléfono'),
            ('D1', 'carrera', 'Carrera o profesión'),
            ('E1', 'grupo', 'Grupo (IMPORTANTE)')
        ]
        
        # Aplicar estilos a encabezados
        for cell_ref, encabezado, descripcion in encabezados:
            cell = ws[cell_ref]
            cell.value = encabezado
            
            if encabezado == 'grupo':
                # Destacar la columna de grupos
                cell.font = grupo_font
                cell.fill = grupo_fill
            else:
                cell.font = header_font
                cell.fill = header_fill
                
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos de ejemplo con grupos
        datos = [
            ['Juan', 'Pérez', '51987654321', 'Ingeniería de Sistemas', 'Ingenierías'],
            ['María', 'González', '51912345678', 'Medicina', 'Medicina'],
            ['Carlos', 'López', '51911223344', 'Derecho', 'Derecho'],
            ['Ana', 'Martínez', '51999887766', 'Psicología', 'Psicología'],
            ['Luis', 'Rodríguez', '51955443322', 'Administración', 'Administración'],
            ['Sofia', 'Fernández', '51977665544', 'Arquitectura', 'Ingenierías'],
            ['Diego', 'García', '51933445566', 'Contabilidad', 'Administración'],
            ['Valentina', 'Hernández', '51988990011', 'Enfermería', 'Medicina'],
            ['Miguel', 'Jiménez', '51922334455', 'Marketing', 'Administración'],
            ['Camila', 'Morales', '51966778899', 'Educación', 'Psicología'],
            ['Pedro', 'Sánchez', '51944556677', 'Estudiante de Secundaria', 'H.S'],
            ['Laura', 'Vega', '51955667788', 'Estudiante de Secundaria', 'H.S']
        ]
        
        # Agregar datos con estilos
        for i, fila in enumerate(datos, start=2):
            for j, valor in enumerate(fila, 1):
                cell = ws.cell(row=i, column=j, value=valor)
                cell.border = thin_border
                
                # Destacar la columna de grupos
                if j == 5:  # Columna E (grupo)
                    cell.font = Font(bold=True, color="FF6B35")
        
        # Ajustar ancho de columnas
        column_widths = [15, 15, 15, 25, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Agregar información sobre grupos en la parte inferior
        ws['A12'] = "GRUPOS DISPONIBLES:"
        ws['A12'].font = Font(bold=True, size=12, color="25D366")
        
        grupos_info = [
            ("A13", "H.S", "Estudiantes de High School (Secundaria)"),
            ("A14", "Ingenierías", "Estudiantes y profesionales de ingeniería"),
            ("A15", "Medicina", "Estudiantes y profesionales de medicina"),
            ("A16", "Derecho", "Estudiantes y profesionales de derecho"),
            ("A17", "Administración", "Estudiantes y profesionales de administración"),
            ("A18", "Psicología", "Estudiantes y profesionales de psicología"),
            ("A18", "General", "Grupo por defecto (sin especificar)")
        ]
        
        for cell_ref, grupo, descripcion in grupos_info:
            ws[cell_ref] = f"{grupo}: {descripcion}"
            ws[cell_ref].font = Font(size=10)
        
        # Agregar instrucciones
        ws['A20'] = "INSTRUCCIONES:"
        ws['A20'].font = Font(bold=True, size=12, color="25D366")
        
        instrucciones = [
            "1. Completa todas las filas con los datos de tus contactos",
            "2. En la columna 'grupo', usa EXACTAMENTE uno de los grupos disponibles arriba",
            "3. Los números deben incluir código de país (ej: 51987654321)",
            "4. Guarda el archivo y súbelo en la aplicación",
            "5. Los grupos se asignarán automáticamente a tus contactos"
        ]
        
        for i, instruccion in enumerate(instrucciones, 21):
            ws[f'A{i}'] = instruccion
            ws[f'A{i}'].font = Font(size=10)
        
        # Crear hoja de instrucciones detalladas
        ws_inst = wb.create_sheet("Instrucciones Detalladas")
        
        # Título
        ws_inst['A1'] = "GUÍA COMPLETA PARA IMPORTAR CONTACTOS CON GRUPOS"
        ws_inst['A1'].font = Font(bold=True, size=16, color="25D366")
        
        # Contenido de instrucciones
        contenido = [
            "",
            "PASO 1: PREPARAR TUS DATOS",
            "• Asegúrate de tener la siguiente información para cada contacto:",
            "  - Nombre (obligatorio)",
            "  - Apellido (opcional)",
            "  - Número de teléfono con código de país (obligatorio)",
            "  - Carrera o profesión (opcional)",
            "  - Grupo al que pertenece (opcional pero recomendado)",
            "",
            "PASO 2: USAR LA COLUMNA DE GRUPOS",
            "• La columna 'grupo' es muy importante para organizar tus contactos",
            "• Usa EXACTAMENTE uno de estos nombres de grupo:",
            "  ✓ H.S (High School - Secundaria)",
            "  ✓ Ingenierías", 
            "  ✓ Medicina",
            "  ✓ Derecho",
            "  ✓ Administración",
            "  ✓ Psicología",
            "  ✓ General",
            "",
            "PASO 3: COMPLETAR EL ARCHIVO",
            "• Copia la estructura de las filas de ejemplo",
            "• Reemplaza los datos de ejemplo con tus contactos reales",
            "• Mantén el mismo formato de columnas",
            "• No cambies los nombres de las columnas",
            "",
            "PASO 4: IMPORTAR EN LA APLICACIÓN",
            "• Ve a la pestaña 'Importar' en la aplicación",
            "• Selecciona este archivo Excel",
            "• Los grupos se asignarán automáticamente",
            "• Podrás enviar mensajes por grupo específico",
            "",
            "BENEFICIOS DE USAR GRUPOS:",
            "• Organización mejorada de contactos",
            "• Envío de mensajes dirigidos por área profesional",
            "• Filtrado eficiente de contactos",
            "• Gestión centralizada de diferentes audiencias",
            "",
            "NOTAS IMPORTANTES:",
            "• Los números deben tener formato internacional (ej: 51987654321)",
            "• Los nombres de grupos deben coincidir exactamente",
            "• Puedes dejar la columna grupo vacía (se asignará 'General')",
            "• El sistema validará los datos antes de importar"
        ]
        
        for i, texto in enumerate(contenido, 2):
            cell = ws_inst[f'A{i}']
            cell.value = texto
            if texto.startswith(('PASO', 'BENEFICIOS', 'NOTAS')):
                cell.font = Font(bold=True, size=12, color="25D366")
            elif texto.startswith('✓'):
                cell.font = Font(bold=True, color="25D366")
            else:
                cell.font = Font(size=10)
        
        # Ajustar ancho de columna en hoja de instrucciones
        ws_inst.column_dimensions['A'].width = 80
        
        # AÑADIR VALIDACIÓN DE DATOS (DESPLEGABLE) PARA LA COLUMNA DE GRUPOS
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            
            # Lista de grupos válidos
            grupos_validos = ['H.S', 'Ingenierías', 'Medicina', 'Derecho', 'Administración', 'Psicología', 'General']
            
            # Crear validación de datos para la columna E (grupos)
            dv = DataValidation(type="list", formula1=f'"{",".join(grupos_validos)}"', allow_blank=True)
            dv.error = 'Selecciona un grupo válido de la lista'
            dv.errorTitle = 'Grupo Inválido'
            dv.prompt = 'Selecciona un grupo de la lista desplegable'
            dv.promptTitle = 'Seleccionar Grupo'
            
            # Aplicar validación a la columna E desde la fila 2 hasta la 1000
            dv.add('E2:E1000')
            ws.add_data_validation(dv)
            
            # Añadir comentario a la columna de grupos (solo si openpyxl está disponible)
            try:
                from openpyxl.comments import Comment
                comment = Comment('Selecciona un grupo de la lista desplegable. Grupos disponibles: H.S, Ingenierías, Medicina, Derecho, Administración, Psicología, General', 'Sistema')
                ws['E1'].comment = comment
            except Exception as e:
                print(f"Advertencia: No se pudo añadir comentario: {e}")
            
        except Exception as e:
            # Si hay error con la validación, continuar sin ella
            print(f"Advertencia: No se pudo añadir validación de datos: {e}")
            # Añadir comentario simple (solo si openpyxl está disponible)
            try:
                from openpyxl.comments import Comment
                comment = Comment('Grupos disponibles: H.S, Ingenierías, Medicina, Derecho, Administración, Psicología, General', 'Sistema')
                ws['E1'].comment = comment
            except Exception as e:
                print(f"Advertencia: No se pudo añadir comentario simple: {e}")
        
        # Crear archivo en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='plantilla_contactos_con_grupos.xlsx'
        )
    except Exception as e:
        logger.error(f"Error creando plantilla Excel: {e}")
        return jsonify({'success': False, 'message': f'Error creando plantilla: {str(e)}'})

@app.route('/api/estadisticas')
def api_estadisticas():
    """API para obtener estadísticas"""
    return jsonify(get_estadisticas())

# APIs de integración del chatbot eliminadas

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5001)